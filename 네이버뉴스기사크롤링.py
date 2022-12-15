from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from PyQt5.QtWidgets import *
import openpyxl
from PyQt5 import uic
import sys
import time
import os
import re
import requests
from bs4 import BeautifulSoup

# 크롬 드라이버 자동 업데이트
from webdriver_manager.chrome import ChromeDriverManager


UI_PATH = "네이버뉴스기사크롤링.ui"

class MainDialog(QDialog):
    def __init__(self):
        QDialog.__init__(self,None)
        uic.loadUi(UI_PATH,self)
        self.crawling.clicked.connect(self.crawling_start)
        self.end_btn.clicked.connect(self.end)
    def crawling_start(self):
        index = 2
        wb = openpyxl.Workbook()

        ws = wb.active
        ws['A1'] = '뉴스링크'
        ws['B1'] = '뉴스본문'

         # 브라우저 꺼짐 방지
        chrome_options = Options()
        chrome_options.add_experimental_option("detach",True)

        # 불필요한 에러 메세지 없애기
        chrome_options.add_experimental_option("excludeSwitches",['enable-logging'])

        service = Service(executable_path=ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service,options=chrome_options)

        for page in range(1,149):
            request_headers = { 
                'User-Agent' : ('Mozilla/5.0 (Windows NT 10.0;Win64; x64)\
                AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98\
                Safari/537.36'), 
            } 
            driver.get(f"https://news.naver.com/main/main.naver?mode=LSD&mid=shm&sid1=100#&date=%2000:00:00&page={page}")
            # 스크롤 전 높이
            before_h = driver.execute_script("return window.scrollY")
            while True:
                # 맨 아래로 스크롤을 내린다.
                driver.find_element(By.CSS_SELECTOR,"body").send_keys(Keys.END)

                # 스크롤 사이 로딩 시간
                time.sleep(1)

                # 스크롤 후 높이
                after_h = driver.execute_script("return window.scrollY")

                if after_h == before_h:
                    break
                before_h = after_h
            items = driver.find_elements(By.CSS_SELECTOR,".type06_headline > li")
            for item in items:
                link = item.find_element(By.CSS_SELECTOR,".type06_headline > li > dl > dt > a").get_attribute('href')
                response = requests.get(link,headers=request_headers)
                html = response.text
                news = BeautifulSoup(html,'html.parser')
                contents = str(news.select("div#dic_area"))
                contents = re.sub('(<([^>]+)>)', '', contents)
                self.textBrowser.append(f"기사링크 : {link} \n")
                ws[f"A{index}"] = link
                ws[f"B{index}"] = contents
                wb.save('네이버뉴스_정치.xlsx')
                index += 1
                QApplication.processEvents()
    def end(self):
        sys.exit()
        


QApplication.setStyle("fusion")
app = QApplication(sys.argv)
main_dialog = MainDialog()
main_dialog.show()

sys.exit(app.exec_())
